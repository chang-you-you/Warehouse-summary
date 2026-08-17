from flask import Flask, render_template, request, send_file
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Border, Alignment, Protection
from datetime import datetime
import io
import os
import re

app = Flask(__name__)

# ============================================================
# 数据配置
# ============================================================

STYLE_MAP = {
    '少女A': ('少女', 'A青春靓丽'),
    '少女B': ('少女', 'B街头时尚'),
    '少女C': ('少女', 'C浪漫少女'),
    '少女D': ('少女', 'D甜酷校园'),
    '少淑A': ('少淑', 'A 首尔风尚'),
    '少淑B': ('少淑', 'B轻熟少女'),
    '少淑C': ('少淑', 'C日系'),
    '少淑D': ('少淑', 'D自然文艺'),
    '少淑E': ('少淑', 'E梦幻森林'),
    '中淑A': ('中淑', 'A日常通勤'),
    '中淑B': ('中淑', 'B魅力女人'),
    '中青A': ('中青', 'A 都市丽人'),
    '中青B': ('中青', 'B假日休闲'),
    '贵妇A': ('贵妇', 'A高贵典雅'),
    '贵妇B': ('贵妇', 'B婉约伊人'),
    '贵妇C': ('贵妇', 'C森韵雅序'),
    '老年A': ('老年', 'A都市雅致'),
    '老年B': ('老年', 'B古风婉韵'),
}

CATEGORIES = {
    '少女A': ['卫衣', 'T恤', '打底衫', '衬衫', '毛衣', '毛衣打底', '毛衣外套', '外套', '马甲', '两件套', '套裤', '套裙'],
    '少女B': ['卫衣', 'T恤', '打底衫', '衬衫', '毛衣', '毛衣打底', '毛衣外套', '外套', '马甲', '两件套', '套裤', '套裙', '小衫'],
    '少女C': ['卫衣', 'T恤', '打底衫', '衬衫', '毛衣', '毛衣打底', '毛衣外套', '外套', '马甲', '两件套', '套裤', '套裙', '小衫'],
    '少女D': ['卫衣', 'T恤', '打底衫', '衬衫', '毛衣', '毛衣打底', '毛衣外套', '外套', '马甲', '两件套', '套裤', '套裙'],
    '少淑A': ['卫衣', '小衫', '打底衫', 'T恤', '衬衫', '针织衫', '毛衣', '毛衣打底', '毛衣外套', '马甲', '两件套', '外套', '小香风', '羊绒大衣', '风衣', '套裤', '套裙', '皮衣'],
    '少淑B': ['卫衣', '小衫', '打底衫', 'T恤', '衬衫', '针织衫', '毛衣', '毛衣打底', '毛衣外套', '马甲', '两件套', '外套', '小香风', '羊绒大衣', '风衣', '套裤', '套裙', '披肩'],
    '少淑C': ['卫衣', '小衫', '打底衫', 'T恤', '衬衫', '针织衫', '毛衣', '毛衣打底', '毛衣外套', '马甲', '两件套', '外套', '小香风', '羊绒大衣', '风衣', '套裤', '套裙'],
    '少淑D': ['卫衣', '小衫', '打底衫', 'T恤', '衬衫', '针织衫', '毛衣', '毛衣打底', '毛衣外套', '马甲', '两件套', '外套', '小香风', '羊绒大衣', '风衣', '套裤', '套裙'],
    '少淑E': ['卫衣', '小衫', '打底衫', 'T恤', '衬衫', '针织衫', '毛衣', '毛衣打底', '毛衣外套', '马甲', '两件套', '外套', '小香风', '羊绒大衣', '风衣', '套裤', '套裙'],
    '中淑A': ['卫衣', '小衫', '打底衫', 'T恤', '衬衫', '针织衫', '毛衣', '毛衣打底', '毛衣外套', '马甲', '两件套', '毛呢', '外套', '小香风', '羊绒大衣', '风衣', '套裤', '套裙', '皮衣'],
    '中淑B': ['卫衣', '小衫', '打底衫', 'T恤', '衬衫', '针织衫', '毛衣', '毛衣打底', '毛衣外套', '马甲', '两件套', '毛呢', '外套', '小香风', '羊绒大衣', '风衣', '套裤', '套裙'],
    '中青A': ['小衫', '打底衫', 'T恤', '衬衫', '针织衫', '毛衣', '毛衣打底', '毛衣外套', '马甲', '毛呢', '外套', '小香风', '羊绒大衣', '风衣', '套裤', '套裙'],
    '中青B': ['小衫', '打底衫', 'T恤', '衬衫', '针织衫', '毛衣', '毛衣打底', '毛衣外套', '马甲', '毛呢', '外套', '小香风', '羊绒大衣', '风衣', '套裤', '套裙'],
    '贵妇A': ['小衫', '打底衫', 'T恤', '衬衫', '针织衫', '毛衣', '毛衣打底', '毛衣外套', '马甲', '毛呢', '外套', '小香风', '羊绒大衣', '风衣', '套裤', '套裙'],
    '贵妇B': ['小衫', '打底衫', 'T恤', '衬衫', '针织衫', '毛衣', '毛衣打底', '毛衣外套', '马甲', '毛呢', '外套', '小香风', '羊绒大衣', '风衣', '套裤', '套裙'],
    '贵妇C': ['小衫', '打底衫', 'T恤', '衬衫', '针织衫', '毛衣', '毛衣打底', '毛衣外套', '马甲', '毛呢', '外套', '小香风', '羊绒大衣', '风衣', '套裤', '套裙'],
    '老年A': ['小衫', '打底衫', 'T恤', '衬衫', '针织衫', '毛衣', '毛衣打底', '毛衣外套', '马甲', '毛呢', '外套', '小香风', '羊绒大衣', '风衣', '套裤', '套裙'],
    '老年B': ['小衫', '打底衫', '衬衫', '针织衫', '毛衣', '毛衣打底', '毛衣外套', '毛呢', '外套', '羊绒大衣', '风衣', '套裤', '套裙'],
}

LOCATIONS = ['西柳', '哈', '郑州']

# ============================================================
# 核心处理函数
# ============================================================

def clean_text(text):
    if text is None:
        return ''
    text = str(text).replace('\n', '').replace('\r', '').replace(' ', '').replace('<br>', '')
    text = text.strip()
    return text


def find_row_by_style_category(sheet, style_name, category_name):
    style_clean = clean_text(style_name)
    category_clean = clean_text(category_name)

    style_row = None
    for row in range(3, sheet.max_row + 1):
        style_cell = clean_text(sheet.cell(row=row, column=2).value)
        if style_cell == style_clean:
            style_row = row
            break

    if style_row is None:
        return None

    end_row = sheet.max_row + 1
    for row in range(style_row + 1, sheet.max_row + 1):
        style_cell = clean_text(sheet.cell(row=row, column=2).value)
        if style_cell != '':
            end_row = row
            break

    for row in range(style_row, end_row):
        category_cell = clean_text(sheet.cell(row=row, column=3).value)
        if category_cell == category_clean:
            return row

    return None


def find_column_by_location_month_day(sheet, location, month, day):
    """查找已存在的【地点+月份+日期】列"""
    for col in range(4, sheet.max_column + 1):
        header = sheet.cell(row=2, column=col).value
        if header is None:
            continue
        header_clean = clean_text(str(header))
        pattern = rf'{re.escape(location)}{month}月(\d+)'
        match = re.search(pattern, header_clean)
        if match:
            header_day = int(match.group(1))
            if header_day == day:
                return col
    return None


def find_first_empty_column(sheet):
    """
    找到第一个空列（从第4列开始扫描）
    空列定义为：第2行没有标题的列
    """
    for col in range(4, sheet.max_column + 2):  # +2 是为了在满列时能创建新列
        header = sheet.cell(row=2, column=col).value
        if header is None or clean_text(str(header)) == '':
            return col
    return sheet.max_column + 1


def copy_cell_style_safe(source_cell, target_cell):
    try:
        if source_cell is None or target_cell is None:
            return
        
        if source_cell.font:
            try:
                target_cell.font = Font(
                    name=source_cell.font.name,
                    size=source_cell.font.size,
                    bold=source_cell.font.bold,
                    italic=source_cell.font.italic,
                    vertAlign=source_cell.font.vertAlign,
                    underline=source_cell.font.underline,
                    strike=source_cell.font.strike,
                    color=source_cell.font.color
                )
            except:
                pass
        
        if source_cell.border:
            try:
                target_cell.border = Border(
                    left=source_cell.border.left,
                    right=source_cell.border.right,
                    top=source_cell.border.top,
                    bottom=source_cell.border.bottom,
                    diagonal=source_cell.border.diagonal,
                    diagonal_direction=source_cell.border.diagonal_direction,
                    outline=source_cell.border.outline
                )
            except:
                pass
        
        if source_cell.fill:
            try:
                target_cell.fill = PatternFill(
                    fill_type=source_cell.fill.fill_type,
                    start_color=source_cell.fill.start_color,
                    end_color=source_cell.fill.end_color
                )
            except:
                pass
        
        if source_cell.number_format:
            try:
                target_cell.number_format = source_cell.number_format
            except:
                pass
        
        if source_cell.alignment:
            try:
                target_cell.alignment = Alignment(
                    horizontal=source_cell.alignment.horizontal,
                    vertical=source_cell.alignment.vertical,
                    text_rotation=source_cell.alignment.text_rotation,
                    wrap_text=source_cell.alignment.wrap_text,
                    shrink_to_fit=source_cell.alignment.shrink_to_fit,
                    indent=source_cell.alignment.indent
                )
            except:
                pass
    except:
        pass


def add_or_find_column(sheet, location, month, day):
    """
    新逻辑：
    1. 先查找是否已存在「地点+月份+日期」的列
    2. 如果存在，返回该列
    3. 如果不存在，找到第一个空列，设置标题，返回该列
    """
    # 1. 先查找是否已存在
    existing_col = find_column_by_location_month_day(sheet, location, month, day)
    if existing_col is not None:
        return existing_col
    
    # 2. 不存在，找到第一个空列
    new_col = find_first_empty_column(sheet)
    
    # 3. 设置标题
    new_header = f"{location}{month}月{day}"
    sheet.cell(row=2, column=new_col, value=new_header)
    
    # 4. 复制样式（使用D列作为样式源）
    for row in range(3, sheet.max_row + 1):
        source_cell = sheet.cell(row=row, column=4)
        target_cell = sheet.cell(row=row, column=new_col)
        copy_cell_style_safe(source_cell, target_cell)
    
    return new_col


def update_sum_formulas(sheet):
    sum_col = None
    for col in range(4, sheet.max_column + 1):
        header = sheet.cell(row=2, column=col).value
        if header and clean_text(str(header)) == '总数':
            sum_col = col
            break
    
    if sum_col is None:
        return
    
    data_end_col = sum_col - 1
    data_end_col_letter = get_column_letter(data_end_col)
    
    for row in range(3, sheet.max_row + 1):
        cell = sheet.cell(row=row, column=sum_col)
        if cell.data_type == 'f' and cell.value and isinstance(cell.value, str) and cell.value.startswith('=SUM('):
            match = re.search(r'=SUM\(([A-Z]+)(\d+):([A-Z]+)(\d+)\)', cell.value)
            if match:
                start_col_letter = match.group(1)
                start_row = match.group(2)
                end_row = match.group(4)
                new_formula = f"=SUM({start_col_letter}{start_row}:{data_end_col_letter}{end_row})"
                cell.value = new_formula


# ============================================================
# Flask 路由
# ============================================================

@app.route('/')
def index():
    style_list = sorted(STYLE_MAP.keys())
    return render_template('index.html', 
                         styles=style_list, 
                         categories=CATEGORIES,
                         locations=LOCATIONS,
                         months=range(1, 13),
                         days=range(1, 32))


@app.route('/get_categories/<style_key>')
def get_categories(style_key):
    return {'categories': CATEGORIES.get(style_key, [])}


@app.route('/upload', methods=['POST'])
def upload_files():
    # ====== 1. 获取上传的Excel文件 ======
    excel_file = request.files['excel_file']
    if not excel_file:
        return "❌ 请上传Excel文件", 400

    # ====== 2. 解析表单数据 ======
    daily_data = []
    style_keys = request.form.getlist('style[]')
    categories = request.form.getlist('category[]')
    values = request.form.getlist('value[]')
    locations = request.form.getlist('location[]')
    months = request.form.getlist('month[]')
    days = request.form.getlist('day[]')

    for i in range(len(style_keys)):
        if i >= len(categories) or i >= len(values) or i >= len(locations) or i >= len(months) or i >= len(days):
            continue
        
        style_key = style_keys[i].strip()
        category = categories[i].strip()
        location = locations[i].strip()
        try:
            month = int(months[i].strip())
            day = int(days[i].strip())
            value = int(values[i].strip())
            if month < 1 or month > 12 or day < 1 or day > 31 or value < 1 or value > 50:
                continue
        except ValueError:
            continue

        if style_key not in STYLE_MAP:
            continue
        if location not in LOCATIONS:
            continue

        sheet_name, style_name = STYLE_MAP[style_key]
        daily_data.append((sheet_name, style_name, category, location, month, day, value))

    if not daily_data:
        return "❌ 未输入有效数据", 400

    # ====== 3. 加载Excel并写入数据 ======
    wb = load_workbook(io.BytesIO(excel_file.read()))

    success_count = 0
    fail_count = 0
    fail_details = []
    created_columns = []

    for sheet_name, style_name, category, location, month, day, value in daily_data:
        if sheet_name not in wb.sheetnames:
            fail_count += 1
            fail_details.append(f"{sheet_name}：Sheet不存在")
            continue
        
        sheet = wb[sheet_name]
        row = find_row_by_style_category(sheet, style_name, category)
        if row is None:
            fail_count += 1
            fail_details.append(f"{sheet_name}/{style_name}->{category}：未找到匹配行")
            continue
        
        col = add_or_find_column(sheet, location, month, day)
        if col is None:
            fail_count += 1
            fail_details.append(f"{sheet_name}/{style_name}->{category}：无法创建列")
            continue
        else:
            # 检查是否是新创建的列（标题是我们刚设置的）
            header = sheet.cell(row=2, column=col).value
            if header and f"{location}{month}月{day}" == str(header):
                # 检查该列是否有数据（判断是否真正新建）
                has_data = False
                for r in range(3, sheet.max_row + 1):
                    if sheet.cell(row=r, column=col).value is not None:
                        has_data = True
                        break
                if not has_data:
                    col_letter = get_column_letter(col)
                    created_columns.append(f"{location} {month}月{day}日 (列{col_letter})")
        
        sheet.cell(row=row, column=col, value=value)
        success_count += 1
        print(f"✅ [{sheet_name}] {style_name} -> {category} = {value} (行{row}, 列{get_column_letter(col)})")

    # ====== 4. 更新所有Sheet的SUM公式 ======
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        update_sum_formulas(sheet)

    # ====== 5. 保存文件到内存 ======
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    # ====== 6. 生成固定文件名 ======
    today_str = datetime.now().strftime('%Y%m%d')
    download_name = f"{today_str}.xlsx"

    # ====== 7. 打印结果日志 ======
    result_msg = f"✅ 成功写入 {success_count} 条数据"
    if created_columns:
        result_msg += f"\n🆕 新增列：{', '.join(created_columns)}"
    if fail_count > 0:
        result_msg += f"\n⚠️ 有 {fail_count} 条数据写入失败"
        for detail in fail_details:
            result_msg += f"\n  ❌ {detail}"

    print(result_msg)

    # ====== 8. 返回文件下载 ======
    return send_file(
        output,
        download_name=download_name,
        as_attachment=True,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000)